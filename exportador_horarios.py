#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Módulo de Exportación de Horarios
Exporta horarios a Excel y HTML de forma profesional
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from typing import List, Dict
import json
from datetime import datetime
from config_horarios import DIAS_SEMANA, HORAS_INICIO, COLORES_MATERIAS_PASTEL


class ExportadorHorarios:
    """Exporta horarios a diferentes formatos"""
    
    def __init__(self):
        self.dias = DIAS_SEMANA
        self.horas = HORAS_INICIO
        # Paleta de colores para materias (colores pasteles profesionales)
        self.colores_materias = COLORES_MATERIAS_PASTEL
        self.mapa_colores = {}
    
    def _obtener_color_materia(self, materia: str) -> str:
        """Asigna un color consistente a cada materia"""
        if materia not in self.mapa_colores:
            idx = len(self.mapa_colores) % len(self.colores_materias)
            self.mapa_colores[materia] = self.colores_materias[idx]
        return self.mapa_colores[materia]
    
    def _agrupar_asignaciones_por_grupo(self, asignaciones: List[Dict], 
                                        eventos: List[Dict]) -> Dict[str, List[Dict]]:
        """Agrupa las asignaciones por grupo"""
        grupos = {}
        
        for asig in asignaciones:
            evento_id = asig['evento_id']
            if evento_id < len(eventos):
                evento = eventos[evento_id]
                grupo = evento['grupo']
                
                if grupo not in grupos:
                    grupos[grupo] = []
                
                grupos[grupo].append({
                    'materia': evento['materia'],
                    'profesor': evento['profesor'],
                    'horas': evento['horas'],
                    'dia': asig['dia'],
                    'hora': asig['hora'],
                    'timeslot': asig['timeslot']
                })
        
        return grupos
    
    def exportar_excel(self, asignaciones: List[Dict], eventos: List[Dict],
                      metricas: Dict, archivo_salida: str = "horarios_completos.xlsx") -> str:
        """
        Exporta TODOS los horarios a Excel con formato profesional
        
        Args:
            asignaciones: Lista de asignaciones de timeslots
            eventos: Lista de eventos
            metricas: Métricas del algoritmo
            archivo_salida: Nombre del archivo de salida
            
        Returns:
            Ruta del archivo generado
        """
        wb = Workbook()
        wb.remove(wb.active)  # Remover hoja por defecto
        
        # Agrupar por grupo
        grupos = self._agrupar_asignaciones_por_grupo(asignaciones, eventos)
        
        # Crear hoja para cada grupo
        for nombre_grupo in sorted(grupos.keys()):
            self._crear_hoja_grupo(wb, nombre_grupo, grupos[nombre_grupo])
        
        # Crear hoja resumen
        self._crear_hoja_resumen(wb, grupos, metricas)
        
        # Guardar
        wb.save(archivo_salida)
        return archivo_salida
    
    def _crear_hoja_grupo(self, wb: Workbook, nombre_grupo: str, 
                          asignaciones_grupo: List[Dict]):
        """Crea una hoja de Excel para un grupo específico"""
        # Nombre de hoja válido (máx 31 caracteres, sin caracteres especiales)
        nombre_hoja = nombre_grupo.replace("/", "-")[:31]
        ws = wb.create_sheet(title=nombre_hoja)
        
        # Encabezado
        ws.merge_cells('A1:F1')
        cell_titulo = ws['A1']
        cell_titulo.value = f"HORARIO - {nombre_grupo}"
        cell_titulo.font = Font(size=16, bold=True, color="FFFFFF")
        cell_titulo.alignment = Alignment(horizontal='center', vertical='center')
        cell_titulo.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        ws.row_dimensions[1].height = 30
        
        # Encabezados de columnas
        headers = ["HORA"] + self.dias
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col_idx)
            cell.value = header
            cell.font = Font(size=12, bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
        
        ws.row_dimensions[2].height = 25
        
        # Crear matriz de horario
        horario_matriz = {}
        for asig in asignaciones_grupo:
            hora = asig['hora']
            dia = asig['dia']
            key = (hora, dia)
            
            if key not in horario_matriz:
                horario_matriz[key] = []
            
            horario_matriz[key].append({
                'materia': asig['materia'],
                'profesor': asig['profesor']
            })
        
        # Llenar celdas
        for row_idx, hora in enumerate(self.horas, start=3):
            # Columna de hora
            cell_hora = ws.cell(row=row_idx, column=1)
            cell_hora.value = hora
            cell_hora.font = Font(size=10, bold=True)
            cell_hora.alignment = Alignment(horizontal='center', vertical='center')
            cell_hora.fill = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
            
            # Columnas de días
            for col_idx, dia in enumerate(self.dias, start=2):
                cell = ws.cell(row=row_idx, column=col_idx)
                key = (hora, dia)
                
                if key in horario_matriz:
                    clases = horario_matriz[key]
                    texto = "\n\n".join([
                        f"{clase['materia']}\n{clase['profesor']}" 
                        for clase in clases
                    ])
                    cell.value = texto
                    
                    # Color según materia
                    if clases:
                        color = self._obtener_color_materia(clases[0]['materia'])
                        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                    
                    cell.font = Font(size=9, bold=True)
                else:
                    cell.value = ""
                    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                # Bordes
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                cell.border = thin_border
            
            ws.row_dimensions[row_idx].height = 60
        
        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 12
        for col in ['B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = 25
    
    def _crear_hoja_resumen(self, wb: Workbook, grupos: Dict, metricas: Dict):
        """Crea hoja resumen con estadísticas"""
        ws = wb.create_sheet(title="📊 RESUMEN", index=0)
        
        # Título
        ws.merge_cells('A1:D1')
        cell_titulo = ws['A1']
        cell_titulo.value = "RESUMEN GENERAL DE HORARIOS"
        cell_titulo.font = Font(size=16, bold=True, color="FFFFFF")
        cell_titulo.alignment = Alignment(horizontal='center', vertical='center')
        cell_titulo.fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
        ws.row_dimensions[1].height = 35
        
        # Fecha de generación
        ws['A3'] = "Fecha de generación:"
        ws['B3'] = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        ws['A3'].font = Font(bold=True)
        
        # Métricas
        ws['A5'] = "MÉTRICAS DEL ALGORITMO"
        ws['A5'].font = Font(size=12, bold=True, color="FFFFFF")
        ws['A5'].fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
        ws.merge_cells('A5:D5')
        
        metricas_lista = [
            ("Tiempo de ejecución (ms)", metricas.get('tiempo_ejecucion_ms', 0)),
            ("Iteraciones", metricas.get('iteraciones', 0)),
            ("Colores (timeslots) usados", metricas.get('colores_usados', 0)),
            ("Conflictos detectados", metricas.get('conflictos_totales', 0)),
            ("Penalización por huecos", metricas.get('penalizacion_huecos', 0)),
            ("Calidad de solución (%)", f"{metricas.get('calidad_solucion', 0):.2f}")
        ]
        
        row = 6
        for metrica, valor in metricas_lista:
            ws.cell(row=row, column=1).value = metrica
            ws.cell(row=row, column=2).value = valor
            ws.cell(row=row, column=1).font = Font(bold=True)
            row += 1
        
        # Estadísticas por grupo
        ws.cell(row=row+1, column=1).value = "ESTADÍSTICAS POR GRUPO"
        ws.cell(row=row+1, column=1).font = Font(size=12, bold=True, color="FFFFFF")
        ws.cell(row=row+1, column=1).fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
        ws.merge_cells(f'A{row+1}:D{row+1}')
        
        # Headers
        row = row + 2
        headers_grupos = ["Grupo", "Materias", "Horas Totales", "Timeslots Ocupados"]
        for col_idx, header in enumerate(headers_grupos, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="95A5A6", end_color="95A5A6", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Datos de grupos
        row += 1
        for nombre_grupo in sorted(grupos.keys()):
            asignaciones = grupos[nombre_grupo]
            materias_unicas = len(set(asig['materia'] for asig in asignaciones))
            horas_totales = sum(asig['horas'] for asig in asignaciones)
            timeslots = len(set(asig['timeslot'] for asig in asignaciones))
            
            ws.cell(row=row, column=1).value = nombre_grupo
            ws.cell(row=row, column=2).value = materias_unicas
            ws.cell(row=row, column=3).value = horas_totales
            ws.cell(row=row, column=4).value = timeslots
            
            # Alineación
            for col in range(1, 5):
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
            
            row += 1
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
    
    def exportar_html(self, asignaciones: List[Dict], eventos: List[Dict],
                     metricas: Dict, info_grafo: Dict,
                     archivo_salida: str = "horarios_completos.html") -> str:
        """
        Exporta TODOS los horarios a HTML con diseño responsive
        
        Args:
            asignaciones: Lista de asignaciones
            eventos: Lista de eventos
            metricas: Métricas del algoritmo
            info_grafo: Información del grafo
            archivo_salida: Nombre del archivo de salida
            
        Returns:
            Ruta del archivo generado
        """
        grupos = self._agrupar_asignaciones_por_grupo(asignaciones, eventos)
        
        html = self._generar_html_header()
        html += self._generar_html_resumen(metricas, info_grafo, len(grupos))
        
        # Generar tabla para cada grupo
        for nombre_grupo in sorted(grupos.keys()):
            html += self._generar_html_grupo(nombre_grupo, grupos[nombre_grupo])
        
        html += self._generar_html_footer()
        
        # Guardar archivo
        with open(archivo_salida, 'w', encoding='utf-8') as f:
            f.write(html)
        
        return archivo_salida
    
    def _generar_html_header(self) -> str:
        """Genera el header HTML con estilos"""
        return """<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Horarios Universitarios - Sistema ITI</title>
    <style>
        * {
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }
        
        body {
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            padding: 20px;
            color: #333;
        }
        
        .container {
            max-width: 1400px;
            margin: 0 auto;
            background: white;
            border-radius: 15px;
            box-shadow: 0 10px 40px rgba(0,0,0,0.3);
            padding: 30px;
        }
        
        h1 {
            text-align: center;
            color: #667eea;
            margin-bottom: 10px;
            font-size: 2.5em;
        }
        
        .subtitle {
            text-align: center;
            color: #666;
            margin-bottom: 30px;
            font-size: 1.1em;
        }
        
        .metrics {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 20px;
            margin-bottom: 40px;
        }
        
        .metric-card {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 20px;
            border-radius: 10px;
            text-align: center;
            box-shadow: 0 5px 15px rgba(0,0,0,0.2);
        }
        
        .metric-card h3 {
            font-size: 0.9em;
            margin-bottom: 10px;
            opacity: 0.9;
        }
        
        .metric-card .value {
            font-size: 2em;
            font-weight: bold;
        }
        
        .grupo-section {
            margin-bottom: 50px;
            page-break-inside: avoid;
        }
        
        .grupo-title {
            background: linear-gradient(90deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 15px 25px;
            border-radius: 10px;
            margin-bottom: 20px;
            font-size: 1.5em;
            box-shadow: 0 5px 15px rgba(0,0,0,0.2);
        }
        
        table {
            width: 100%;
            border-collapse: collapse;
            margin-bottom: 30px;
            background: white;
            border-radius: 10px;
            overflow: hidden;
            box-shadow: 0 5px 15px rgba(0,0,0,0.1);
        }
        
        th {
            background: #34495e;
            color: white;
            padding: 15px;
            text-align: center;
            font-weight: 600;
        }
        
        td {
            padding: 12px;
            border: 1px solid #ddd;
            text-align: center;
            vertical-align: middle;
            min-height: 60px;
        }
        
        td.hora {
            background: #ecf0f1;
            font-weight: bold;
            color: #2c3e50;
        }
        
        td.ocupado {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            font-weight: 500;
            font-size: 0.9em;
        }
        
        td.vacio {
            background: #f8f9fa;
        }
        
        .materia {
            font-weight: bold;
            margin-bottom: 5px;
        }
        
        .profesor {
            font-size: 0.85em;
            opacity: 0.9;
        }
        
        @media print {
            body {
                background: white;
                padding: 0;
            }
            
            .container {
                box-shadow: none;
            }
            
            .grupo-section {
                page-break-after: always;
            }
        }
        
        @media (max-width: 768px) {
            h1 {
                font-size: 1.8em;
            }
            
            .metrics {
                grid-template-columns: 1fr;
            }
            
            table {
                font-size: 0.8em;
            }
        }
    </style>
</head>
<body>
    <div class="container">
        <h1>📅 Sistema de Horarios Universitarios</h1>
        <div class="subtitle">Universidad Politécnica de Victoria - Ingeniería en TI</div>
"""
    
    def _generar_html_resumen(self, metricas: Dict, info_grafo: Dict, num_grupos: int) -> str:
        """Genera la sección de resumen con métricas"""
        return f"""
        <div class="metrics">
            <div class="metric-card">
                <h3>Grupos Generados</h3>
                <div class="value">{num_grupos}</div>
            </div>
            <div class="metric-card">
                <h3>Timeslots Usados</h3>
                <div class="value">{metricas.get('colores_usados', 0)}</div>
            </div>
            <div class="metric-card">
                <h3>Calidad</h3>
                <div class="value">{metricas.get('calidad_solucion', 0):.1f}%</div>
            </div>
            <div class="metric-card">
                <h3>Conflictos</h3>
                <div class="value">{metricas.get('conflictos_totales', 0)}</div>
            </div>
            <div class="metric-card">
                <h3>Eventos (Nodos)</h3>
                <div class="value">{info_grafo.get('nodos', 0)}</div>
            </div>
            <div class="metric-card">
                <h3>Tiempo (ms)</h3>
                <div class="value">{metricas.get('tiempo_ejecucion_ms', 0):.0f}</div>
            </div>
        </div>
"""
    
    def _generar_html_grupo(self, nombre_grupo: str, asignaciones: List[Dict]) -> str:
        """Genera la tabla HTML para un grupo"""
        html = f"""
        <div class="grupo-section">
            <div class="grupo-title">{nombre_grupo}</div>
            <table>
                <thead>
                    <tr>
                        <th>HORA</th>
"""
        
        for dia in self.dias:
            html += f"                        <th>{dia}</th>\n"
        
        html += """                    </tr>
                </thead>
                <tbody>
"""
        
        # Crear matriz de horario
        horario_matriz = {}
        for asig in asignaciones:
            hora = asig['hora']
            dia = asig['dia']
            key = (hora, dia)
            
            if key not in horario_matriz:
                horario_matriz[key] = []
            
            horario_matriz[key].append({
                'materia': asig['materia'],
                'profesor': asig['profesor']
            })
        
        # Generar filas
        for hora in self.horas:
            html += f"                    <tr>\n"
            html += f"                        <td class='hora'>{hora}</td>\n"
            
            for dia in self.dias:
                key = (hora, dia)
                if key in horario_matriz:
                    clases = horario_matriz[key]
                    contenido = "<br>".join([
                        f"<div class='materia'>{clase['materia']}</div>"
                        f"<div class='profesor'>{clase['profesor']}</div>"
                        for clase in clases
                    ])
                    html += f"                        <td class='ocupado'>{contenido}</td>\n"
                else:
                    html += f"                        <td class='vacio'></td>\n"
            
            html += "                    </tr>\n"
        
        html += """                </tbody>
            </table>
        </div>
"""
        
        return html
    
    def _generar_html_footer(self) -> str:
        """Genera el footer HTML"""
        return f"""
        <div style="text-align: center; margin-top: 40px; padding-top: 20px; border-top: 2px solid #ddd; color: #666;">
            <p>Generado el {datetime.now().strftime("%d/%m/%Y a las %H:%M:%S")}</p>
            <p>Sistema de Horarios ITI - Graph Coloring Algorithm</p>
        </div>
    </div>
</body>
</html>
"""


if __name__ == "__main__":
    print("Módulo de Exportación de Horarios")
    print("Para usar, importar: from exportador_horarios import ExportadorHorarios")
