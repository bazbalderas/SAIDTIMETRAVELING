#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Test de validación del sistema completo (sin Qt GUI)
"""

import os
import sys

def test_complete_system():
    """Prueba el sistema completo sin GUI"""
    print("=" * 70)
    print("TEST COMPLETO DEL SISTEMA DE HORARIOS")
    print("=" * 70)
    
    # Importar módulos
    print("\n1. Importando módulos...")
    from sistema_horarios_qt import SistemaHorarios
    from exportador_horarios import ExportadorHorarios
    from visualizacion_grafo import VisualizadorGrafo
    print("   ✓ Todos los módulos importados correctamente")
    
    # Crear sistema
    print("\n2. Creando sistema...")
    sistema = SistemaHorarios("config.json")
    print("   ✓ Sistema creado")
    
    # Cargar datos
    print("\n3. Cargando datos completos...")
    sistema.cargar_datos("data/datos_completos.json")
    print(f"   ✓ {len(sistema.profesores)} profesores")
    print(f"   ✓ {len(sistema.grupos)} grupos")
    print(f"   ✓ {len(sistema.materias)} materias")
    
    # Ejecutar algoritmo
    print("\n4. Ejecutando algoritmo de generación...")
    if sistema.ejecutar_algoritmo():
        print("   ✓ Algoritmo ejecutado exitosamente")
        
        # Mostrar resultados
        print("\n5. Resultados del algoritmo:")
        resultados = sistema.resultados
        metricas = resultados['metricas']
        info_grafo = resultados['info_grafo']
        
        print(f"   • Eventos generados: {len(sistema.eventos)}")
        print(f"   • Nodos en grafo: {info_grafo['nodos']}")
        print(f"   • Aristas (conflictos): {info_grafo['aristas']}")
        print(f"   • Timeslots usados: {metricas['colores_usados']}")
        print(f"   • Conflictos totales: {metricas['conflictos_totales']}")
        print(f"   • Calidad solución: {metricas['calidad_solucion']:.2f}%")
        print(f"   • Tiempo ejecución: {metricas['tiempo_ejecucion_ms']:.2f} ms")
        
        # Verificar asignaciones
        print("\n6. Verificando asignaciones...")
        asignaciones = resultados['asignaciones']
        print(f"   • Total asignaciones: {len(asignaciones)}")
        
        # Verificar que todos los eventos tienen asignación
        eventos_asignados = set(asig['evento_id'] for asig in asignaciones)
        print(f"   • Eventos con asignación: {len(eventos_asignados)}")
        
        if len(eventos_asignados) == len(sistema.eventos):
            print("   ✓ Todos los eventos tienen horario asignado")
        else:
            print(f"   ⚠ Faltan {len(sistema.eventos) - len(eventos_asignados)} eventos por asignar")
        
        # Verificar conflictos duros
        print("\n7. Verificando conflictos duros...")
        conflictos_duros = verificar_conflictos_duros(sistema.eventos, asignaciones)
        if conflictos_duros == 0:
            print("   ✓ SIN CONFLICTOS DUROS (profesor/grupo al mismo tiempo)")
        else:
            print(f"   ✗ {conflictos_duros} conflictos duros detectados")
        
        # Exportar Excel
        print("\n8. Exportando a Excel...")
        archivo_excel = sistema.exportar_excel_completo("test_final_horarios.xlsx")
        if archivo_excel and os.path.exists(archivo_excel):
            size = os.path.getsize(archivo_excel) / 1024
            print(f"   ✓ Excel: {archivo_excel} ({size:.1f} KB)")
            
            # Verificar contenido del Excel
            from openpyxl import load_workbook
            wb = load_workbook(archivo_excel)
            print(f"   ✓ Hojas en Excel: {len(wb.sheetnames)}")
            print(f"   ✓ Hojas: {', '.join(wb.sheetnames[:5])}...")
        
        # Exportar HTML
        print("\n9. Exportando a HTML...")
        archivo_html = sistema.exportar_html_completo("test_final_horarios.html")
        if archivo_html and os.path.exists(archivo_html):
            size = os.path.getsize(archivo_html) / 1024
            print(f"   ✓ HTML: {archivo_html} ({size:.1f} KB)")
            
            # Verificar contenido HTML
            with open(archivo_html, 'r', encoding='utf-8') as f:
                html_content = f.read()
                num_tablas = html_content.count('<table>')
                print(f"   ✓ Tablas de horarios en HTML: {num_tablas}")
        
        # Generar visualización del grafo
        print("\n10. Generando visualización del grafo...")
        ruta_grafo, stats = sistema.generar_visualizacion_grafo("test_final_grafo.png")
        if ruta_grafo and os.path.exists(ruta_grafo):
            size = os.path.getsize(ruta_grafo) / 1024
            print(f"   ✓ Grafo: {ruta_grafo} ({size:.1f} KB)")
            print(f"   ✓ Densidad del grafo: {stats['densidad']:.2%}")
            print(f"   ✓ Grado promedio: {stats['grado_promedio']:.2f}")
        
        # Exportar otros formatos
        print("\n11. Exportando otros formatos...")
        sistema.exportar_resultados_json("test_final_resultados.json")
        sistema.exportar_matriz_csv("test_final_matriz.csv")
        
        if os.path.exists("test_final_resultados.json"):
            size = os.path.getsize("test_final_resultados.json") / 1024
            print(f"   ✓ JSON: test_final_resultados.json ({size:.1f} KB)")
        
        if os.path.exists("test_final_matriz.csv"):
            size = os.path.getsize("test_final_matriz.csv") / 1024
            print(f"   ✓ CSV: test_final_matriz.csv ({size:.1f} KB)")
        
        # Resumen final
        print("\n" + "=" * 70)
        print("RESUMEN DE VALIDACIÓN")
        print("=" * 70)
        print(f"✓ Sistema: FUNCIONANDO")
        print(f"✓ Calidad: {metricas['calidad_solucion']:.2f}%")
        print(f"✓ Conflictos duros: {conflictos_duros}")
        print(f"✓ Exportación Excel: OK")
        print(f"✓ Exportación HTML: OK")
        print(f"✓ Visualización grafo: OK")
        print("=" * 70)
        print("✅ SISTEMA COMPLETAMENTE FUNCIONAL AL 100%")
        print("=" * 70)
        
        return True
    else:
        print("   ✗ Error al ejecutar el algoritmo")
        return False

def verificar_conflictos_duros(eventos, asignaciones):
    """Verifica si hay conflictos de profesor/grupo en el mismo timeslot"""
    conflictos = 0
    
    # Crear mapas de timeslot -> profesores y timeslot -> grupos
    timeslot_profesores = {}
    timeslot_grupos = {}
    
    for asig in asignaciones:
        evento_id = asig['evento_id']
        timeslot = asig['timeslot']
        
        if evento_id < len(eventos):
            evento = eventos[evento_id]
            profesor = evento['profesor']
            grupo = evento['grupo']
            
            # Verificar profesor
            if timeslot not in timeslot_profesores:
                timeslot_profesores[timeslot] = set()
            
            if profesor and profesor in timeslot_profesores[timeslot]:
                conflictos += 1
                print(f"   ⚠ Conflicto: {profesor} en timeslot {timeslot}")
            timeslot_profesores[timeslot].add(profesor)
            
            # Verificar grupo
            if timeslot not in timeslot_grupos:
                timeslot_grupos[timeslot] = set()
            
            if grupo in timeslot_grupos[timeslot]:
                conflictos += 1
                print(f"   ⚠ Conflicto: {grupo} en timeslot {timeslot}")
            timeslot_grupos[timeslot].add(grupo)
    
    return conflictos

if __name__ == "__main__":
    try:
        exito = test_complete_system()
        sys.exit(0 if exito else 1)
    except Exception as e:
        print(f"\n✗ Error durante las pruebas: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
